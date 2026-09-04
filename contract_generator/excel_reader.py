"""Чтение карточки компании из книги Excel (.xlsx).

Правила выбора данных (детерминированные и задокументированные):

1. Просматриваются ВСЕ видимые листы книги в порядке их следования.
   Скрытые листы (``hidden``/``veryHidden``) игнорируются.
2. На каждом листе просматриваются первые :data:`MAX_LABEL_COLUMNS` колонок.
   Ячейка считается меткой поля, если её текст сопоставляется с известным
   полем; значением берётся первая непустая ячейка справа от метки в пределах
   :data:`MAX_VALUE_LOOKAHEAD` колонок.
3. Первое найденное значение поля выигрывает. Если то же поле встречается
   ещё раз с ДРУГИМ значением, выдаётся предупреждение о дубликате, значение
   не подменяется молча.

Формат ``.xls`` не поддерживается: openpyxl его не читает.
"""
from __future__ import annotations

import logging
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.utils.datetime import from_excel
from openpyxl.worksheet.worksheet import Worksheet

from . import fields as F
from .models import (
    SOURCE_CARD,
    WARNING,
    CompanyCard,
    FieldValue,
    Issue,
)

logger = logging.getLogger(__name__)

MAX_LABEL_COLUMNS = 4
MAX_VALUE_LOOKAHEAD = 6
MAX_ROWS = 2000


class ExcelReadError(Exception):
    """Карточку компании не удалось прочитать."""


def _is_visible(sheet: Worksheet) -> bool:
    return getattr(sheet, "sheet_state", "visible") == "visible"


def _format_number(value: float) -> str:
    """Число -> строка без экспоненциальной записи и без лишнего ``.0``."""
    if isinstance(value, bool):
        return "да" if value else "нет"
    if isinstance(value, int):
        return str(value)
    if float(value).is_integer():
        # Целые, пришедшие как float, не должны стать "1.2345678901e+11".
        return str(int(round(float(value))))
    return format(float(value), "f").rstrip("0").rstrip(".")


def _cell_to_text(cell: Any, kind: str) -> str:
    """Преобразует значение ячейки в строку с учётом типа поля."""
    value = cell.value if hasattr(cell, "value") else cell
    if value is None:
        return ""

    if isinstance(value, (datetime, date)):
        if kind == F.DATE:
            return _date_to_text(value)
        return _date_to_text(value)

    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if kind == F.DATE:
            # Числовая ячейка с датным форматом — это серийная дата Excel.
            if _looks_like_date_cell(cell):
                try:
                    return _date_to_text(from_excel(value))
                except (ValueError, TypeError, OverflowError):
                    logger.warning("Не удалось преобразовать серийную дату Excel")
            return _format_number(value)
        if kind == F.IDENTIFIER:
            # ИНН/КПП/счета должны остаться строками без экспоненты.
            return _restore_identifier(_format_number(value), cell)
        return _format_number(value)

    text = str(value).strip()
    if kind == F.IDENTIFIER:
        return text.replace(" ", "").replace(" ", "")
    return text


def _looks_like_date_cell(cell: Any) -> bool:
    fmt = str(getattr(cell, "number_format", "") or "").lower()
    return any(token in fmt for token in ("yy", "гг", "mm-dd", "d.m", "dd.mm"))


def _restore_identifier(text: str, cell: Any) -> str:
    """Восстанавливает ведущие нули по числовому формату ячейки.

    Excel хранит «0012345678» как число 12345678 с форматом ``00000000``.
    Если формат содержит достаточно нулей, длина восстанавливается.
    """
    fmt = str(getattr(cell, "number_format", "") or "")
    zeros = fmt.count("0")
    if zeros > len(text) and set(fmt) <= {"0", '"', "_", "-", " ", ";", "@"}:
        return text.rjust(zeros, "0")
    return text


def _date_to_text(value: Any) -> str:
    if isinstance(value, datetime):
        value = value.date()
    if isinstance(value, date):
        return f"{value.day:02d}.{value.month:02d}.{value.year}"
    return str(value)


def _iter_sheet_pairs(sheet: Worksheet) -> List[Tuple[str, Any, str]]:
    """Возвращает список ``(ключ_поля, ячейка_значения, исходная_метка)``."""
    found: List[Tuple[str, Any, str]] = []
    max_col = min(sheet.max_column or 1, MAX_LABEL_COLUMNS + MAX_VALUE_LOOKAHEAD)
    max_row = min(sheet.max_row or 1, MAX_ROWS)

    for row_idx in range(1, max_row + 1):
        for col_idx in range(1, min(max_col, MAX_LABEL_COLUMNS) + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            raw_label = cell.value
            if raw_label is None or not str(raw_label).strip():
                continue
            key = F.resolve_field(raw_label)
            if not key:
                continue
            value_cell = None
            for offset in range(1, MAX_VALUE_LOOKAHEAD + 1):
                candidate = sheet.cell(row=row_idx, column=col_idx + offset)
                if candidate.value is not None and str(candidate.value).strip():
                    value_cell = candidate
                    break
            found.append((key, value_cell, str(raw_label).strip()))
            break  # одна метка на строку
    return found


def _collect_unknown_labels(sheet: Worksheet) -> List[str]:
    """Метки в первой колонке, которые не удалось сопоставить с полем."""
    unknown: List[str] = []
    max_row = min(sheet.max_row or 1, MAX_ROWS)
    for row_idx in range(1, max_row + 1):
        cell = sheet.cell(row=row_idx, column=1)
        raw = cell.value
        if raw is None or not str(raw).strip():
            continue
        text = str(raw).strip()
        if F.resolve_field(text):
            continue
        neighbour = sheet.cell(row=row_idx, column=2).value
        # Считаем меткой только строки вида "что-то: значение".
        if neighbour is not None and str(neighbour).strip() and len(text) <= 80:
            unknown.append(text)
    return unknown


def read_company_card(path: str | Path) -> CompanyCard:
    """Читает карточку компании из ``.xlsx``.

    :raises ExcelReadError: файл отсутствует, повреждён или имеет формат
        ``.xls``, который не поддерживается.
    """
    file_path = Path(path)
    if not file_path.is_file():
        raise ExcelReadError(f"Файл карточки не найден: {file_path}")
    if file_path.suffix.lower() == ".xls":
        raise ExcelReadError(
            "Формат .xls не поддерживается. Сохраните карточку в формате .xlsx "
            "(Excel: «Сохранить как» → «Книга Excel (*.xlsx)»)."
        )

    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as exc:  # openpyxl бросает разнородные исключения
        logger.exception("Не удалось открыть книгу Excel")
        raise ExcelReadError(
            f"Не удалось открыть файл карточки: {exc}"
        ) from exc

    card = CompanyCard(source_path=str(file_path))
    seen_labels: Dict[str, Tuple[str, str, str]] = {}  # key -> (value, sheet, label)

    try:
        visible_sheets = [s for s in workbook.worksheets if _is_visible(s)]
        card.sheet_names = [s.title for s in visible_sheets]
        if not visible_sheets:
            raise ExcelReadError("В книге нет видимых листов.")

        for sheet in visible_sheets:
            for key, value_cell, raw_label in _iter_sheet_pairs(sheet):
                spec = F.FIELDS_BY_KEY[key]
                text = _cell_to_text(value_cell, spec.kind) if value_cell is not None else ""

                if key in seen_labels:
                    prev_value, prev_sheet, prev_label = seen_labels[key]
                    if text and text != prev_value:
                        card.warnings.append(Issue(
                            key,
                            f'Поле «{spec.display}» встречается несколько раз с разными '
                            f'значениями: «{prev_value}» (лист «{prev_sheet}», метка '
                            f'«{prev_label}») и «{text}» (лист «{sheet.title}», метка '
                            f'«{raw_label}»). Использовано первое — проверьте.',
                            WARNING,
                        ))
                    continue

                seen_labels[key] = (text, sheet.title, raw_label)
                card.values[key] = FieldValue(
                    key, text, SOURCE_CARD if text else "empty"
                )

            card.unknown_labels.extend(
                label for label in _collect_unknown_labels(sheet)
                if label not in card.unknown_labels
            )
    finally:
        workbook.close()

    # Поля, которых нет в карточке вовсе, должны существовать как пустые.
    for spec in F.FIELD_SPECS:
        card.values.setdefault(spec.key, FieldValue(spec.key, "", "empty"))

    logger.info(
        "Карточка прочитана: %s, листов=%d, заполненных полей=%d",
        file_path.name,
        len(card.sheet_names),
        sum(1 for v in card.values.values() if not v.is_empty),
    )
    return card


def load_company_data(path: str | Path) -> Dict[str, str]:
    """Совместимость со старым API: возвращает ``{метка Excel: значение}``."""
    card = read_company_card(path)
    result: Dict[str, str] = {}
    for key, value in card.values.items():
        if value.is_empty:
            continue
        spec = F.FIELDS_BY_KEY.get(key)
        if spec and spec.aliases:
            result[spec.aliases[0]] = value.value
    return result
